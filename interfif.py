import datetime
from datetime import timedelta

import tkinter
from tkinter import *
from tkinter import filedialog

import openpyxl
import requests.api
import time
import json.encoder
import sys


def getTypeSIFromStringCell(v):
    r = v[-12:]
    result = ''
    for s in r:
        if s in ['-', '0', '1', '2', '3', '4', '5', '6', '7', '8', '9']:
            result += s
    return result


def getSerialNumberFromStringCell(v):
    r = v.split(',')[-1:][0]
    result = ''
    for s in r:
        if s not in [')', '(', '"']:
            result += s
    return result


# "verification_date": "2020-12-01T00:00:00Z",
def getStrFromDataCell(v):
    r = v.split(',')[1]
    r = r.split('.')[0]
    dd = ''
    for s in r:
        if s not in [')', '(', '"']:
            dd += s
    d = datetime.date(1900, 1, 1)
    d = d + timedelta(int(dd))
    result = d.strftime("%Y-%m-%dT00:00:00Z")
    return result


def getPost(cletB, cletD, cletF):
    u = 'https://fgis.gost.ru/fundmetrology/eapi/vri'
    param = {'year': cletF[:4], 'search': cletB + ' ' + cletD}
    time.sleep(1)
    resp = requests.get(url=u, params=param, verify=False)
    # print(resp.text)
    data = json.loads(resp.text)
    res = data['result']
    if res['count'] == 0:
        return ''
    for item in res['items']:
        if 'ДЮЮ' in item['result_docnum']:
            return item['result_docnum']
    return resp.text


def getShortArshinNumber(fullNumber):
    t = fullNumber.split('/')
    return t[-1]


def PeekFile():
    file = filedialog.askopenfilename(filetypes = (("Excel files","*.xlsx"),("all files","*.*")))
    lblFile.configure(text=file)


def AskFromArshin():
    lblRes.configure(text="Запрашиваю данные из Аршина")
    fileName =  lblFile.cget("text")
    wb = openpyxl.load_workbook(fileName)
    sheet = wb['Лист1']
    cB = 2
    cD = 4
    cF = 6
    cAK = 37
    r = 2
    while r < 65536:
        cletB = sheet.cell(row=r, column=cB).value
        if cletB == None:
            break
        vCellB = getTypeSIFromStringCell(cletB)
        cletD = sheet.cell(row=r, column=cD).value
        vCellD = getSerialNumberFromStringCell(cletD)
        cletF = sheet.cell(row=r, column=cF).value
        vCellF = getStrFromDataCell(cletF)
        sheet.cell(row=r, column=cAK).value = getShortArshinNumber(getPost(vCellB, vCellD, vCellF))
        r += 1
    wb.save(fileName)
    lblRes.configure(text="Данные из Аршина получены")


window = Tk()
window.geometry('500x250')
window.title("InterFIF - загрузка данных из ФГИС Аршин")
lblInfo = Label(window, text="Выберите файл xlsx")
lblInfo.grid(column=0, row=0)
buttonPeek = Button(window, text="Открыть XLSX файл", command=PeekFile)
buttonPeek.grid(column=0, row=1)
lblFile = Label(window, text="Файл не выбран")
lblFile.grid(column=1, row=1)
buttonPeek = Button(window, text="Выполнить запрос из ФГИС Аршин", command=AskFromArshin)
buttonPeek.grid(column=0, row=2)
lblRes = Label(window, text="Запрос не произведён")
lblRes.grid(column=1, row=2)
window.mainloop()






